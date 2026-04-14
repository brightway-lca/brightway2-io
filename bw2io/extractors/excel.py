import os
from pathlib import Path

from openpyxl import cell, load_workbook, workbook


def get_cell_value_handle_error(cell: cell.cell.Cell):
    """
    Retrieve the value of a given cell and handle error types.

    Parameters
    ----------
    cell : openpyxl.cell.cell.Cell
        The cell to get the value from.

    Returns
    --------
    object
        The value of the cell, or None if the cell has an error type.

    Examples
    ---------
    >>> from openpyxl import Workbook
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws["A1"] = "hello"
    >>> assert get_cell_value_handle_error(ws["A1"]) == "hello"
    >>> ws["B1"] = "=1/0"
    >>> assert get_cell_value_handle_error(ws["B1"]) == None
    """
    if cell.data_type == "e":
        # Error type
        return None
    else:
        return cell.value


class ExcelExtractor:
    """
    A class used to extract data from an Excel file.

    Parameters
    ----------
    object : type
        The parent object for the ExcelExtractor class.

    Returns
    -------
    object
        An instance of the class.

    See Also
    --------
    openpyxl.load_workbook : Load a workbook from a file.

    Notes
    -----
    This class requires the openpyxl package to be installed.

    Raises
    ------
    AssertionError
        If the file at 'filepath' does not exist.

    Parameters
    ----------
    filepath : str
        The path to the Excel file.

    Returns
    -------
    list
        A list of tuples containing the name of each sheet in the file and the data from each sheet.

    Examples
    --------
    >>> extractor = ExcelExtractor()
    >>> filepath = 'example.xlsx'
    >>> data = extractor.extract(filepath)
    """

    @staticmethod
    def _normalize_sheet_names(value):
        """Coerce the ``sheet_name`` argument to a list of strings, or ``None``."""
        if value is None:
            return None
        if isinstance(value, str):
            return [value]
        if isinstance(value, (list, tuple)):
            return list(value)
        raise TypeError(
            "sheet_name must be a string, list, or tuple of strings, got {!r}".format(
                type(value).__name__
            )
        )

    @classmethod
    def extract(cls, filepath: Path, sheet_name=None, **kwargs):
        """
        Extract data from an Excel file.

        Parameters
        ----------
        filepath : str
            The path to the Excel file.
        sheet_name : str or list of str or None
            If given, only extract the named sheet(s).  A single sheet name
            may be passed as a string; multiple sheets as a list or tuple.
            ``None`` (the default) extracts all sheets.

        Returns
        -------
        list
            A list of tuples containing the name of each sheet in the file and the data from each sheet.

        Raises
        ------
        AssertionError
            If the file at 'filepath' does not exist.
        ValueError
            If any requested sheet name is not present in the workbook.
        """
        filepath = Path(filepath)
        assert filepath.is_file(), "Can't file file at path {}".format(filepath)
        wb = load_workbook(filepath, data_only=True, read_only=True)
        sheet_names = cls._normalize_sheet_names(sheet_name)
        if sheet_names is None:
            selected = wb.sheetnames
        else:
            missing = [name for name in sheet_names if name not in wb.sheetnames]
            if missing:
                wb.close()
                raise ValueError(
                    "Unknown sheet name(s): {}".format(", ".join(sorted(missing)))
                )
            selected = sheet_names
        data = [(name, cls.extract_sheet(wb, name)) for name in selected]
        wb.close()
        return data

    @classmethod
    def extract_sheet(cls, wb: workbook.Workbook, name: str, strip: bool = True):
        """
        Extract data from a single sheet in an Excel workbook.

        Parameters
        ----------
        wb : openpyxl.workbook.Workbook
            The workbook object with the sheet to extract data from.
        name : str
            The name of the sheet to extract data from.
        strip : bool, optional
            If True, strip whitespace from cell values, by default True.

        Returns
        -------
        list
            A list of lists containing the data from the sheet.

        Notes
        -----
        This method is called by the 'extract' method to extract the data from each sheet in the workbook.

        Examples
        --------
        >>> wb = openpyxl.load_workbook('example.xlsx')
        >>> name = 'Sheet1'
        >>> data = ExcelExtractor.extract_sheet(wb, sheetname)
        """
        ws = wb[name]
        _ = lambda x: x.strip() if (strip and hasattr(x, "strip")) else x
        return [
            [_(get_cell_value_handle_error(cell)) for cell in row] for row in ws.rows
        ]
