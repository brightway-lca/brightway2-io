from ..compatibility import (
    SIMAPRO_BIOSPHERE,
    ECOSPOLD_2_3_BIOSPHERE,
)
from ..units import normalize_units
from bw2data import config, Database, databases, Method, methods, parameters
from bw2data.parameters import Group
from functools import partial
from pathlib import Path
from numbers import Number
from openpyxl import load_workbook
import copy
import csv
import gzip
import json

dirpath = Path(__file__).parent.resolve()


def write_json_file(data, name):
    with open(dirpath / name + ".json", "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)


def get_csv_example_filepath():
    return dirpath / "examples" / "example.csv"


def get_xlsx_example_filepath():
    return dirpath / "examples" / "example.xlsx"


def get_sheet(path, name):
    return load_workbook(path)[name]


# def get_ecoinvent_301_31_migration_data():
#     ws = get_sheet(dirpath / "lci" / "ecoinvent 3.01-3.1.xlsx", "comparison list")
#     deleted_activities = [
#         (ws.cell(row, 0).value, ws.cell(row, 1).value)
#         for row in range(1, ws.max_row)
#         if ws.cell(row, 3).value == "deleted dataset"
#     ]
#     new_activities = [
#         (ws.cell(row, 0).value, ws.cell(row, 1).value)
#         for row in range(1, ws.max_row)
#         if ws.cell(row, 3).value == "new dataset"
#     ]
#     actually_deleted = [x for x in deleted_activities if x not in new_activities]


# def get_ecoinvent_2_301_migration_data():
#     ws = get_sheet(
#         dirpath / "lci" / "ecoinvent 2-3.01.xlsx", "correspondance sheet_corrected"
#     )
#     migration_data = [{
#         '2.2 name': ws.cell(row_index, 2).value,
#         'activity': ws.cell(row_index, 5).value,
#         'product': ws.cell(row_index, 7).value,
#         '2.2 unit': ws.cell(row_index, 10).value,
#         'unit': ws.cell(row_index, 17).value,
#         '2.2 location': ws.cell(row_index, 11).value,
#         'location': ws.cell(row_index, 14).value,
#         'conversion': ws.cell(row_index, 18).value,
#     } for row_index in range(1, ws.max_row)]

#     deleted_activities = [
#         (ws.cell(row, 0).value, ws.cell(row, 1).value)
#         for row in range(1, ws.max_row)
#         if ws.cell(row, 3).value == "deleted dataset"
#     ]
#     new_activities = [
#         (ws.cell(row, 0).value, ws.cell(row, 1).value)
#         for row in range(1, ws.max_row)
#         if ws.cell(row, 3).value == "new dataset"
#     ]
#     actually_deleted = [x for x in deleted_activities if x not in new_activities]


def get_biosphere_2_3_category_migration_data():
    """Get data for 2 -> 3 migration for biosphere flow categories"""
    return {
        "fields": ["categories", "type"],
        "data": [
            ((k, "biosphere"), {"categories": v})  # Exchanges
            for k, v in ECOSPOLD_2_3_BIOSPHERE.items()
        ]
        + [
            ((k, "emission"), {"categories": v})  # Datasets
            for k, v in ECOSPOLD_2_3_BIOSPHERE.items()
        ],
    }


def get_biosphere_2_3_name_migration_data():
    """Get migration data for 2 -> 3 biosphere flow names.

    This migration **must** be applied only after categories have been updated.

    Note that the input data excel sheet is **modified** from the raw data provided by ecoinvent - some biosphere flows which had no equivalent in ecospold2 were mapped using my best judgment. Name changes from 3.1 were also included. Modified cells are marked in **dark orange**.

    Note that not all rows have names in ecoinvent 3. There are a few energy resources that we don't update. For water flows, the categories are updated by a different strategy, and the names don't change, so we just ignore them for now."""

    ws = get_sheet(
        dirpath / "lci" / "ecoinvent elementary flows 2-3.xlsx", "ElementaryExchanges"
    )

    def to_exchange(obj):
        obj[0][3] = u"biosphere"
        return obj

    def strip_unspecified(one, two):
        if two == "unspecified":
            return (one,)
        else:
            return (one, two)

    data = [
        (
            [
                ws.cell(row=row + 1, column=2).value,  # Old name
                # Categories
                strip_unspecified(
                    ws.cell(row=row + 1, column=10).value,
                    ws.cell(row=row + 1, column=11).value,
                ),
                normalize_units(ws.cell(row=row + 1, column=7).value),
                u"emission",  # Unit
            ],
            {"name": ws.cell(row=row + 1, column=9).value},
        )
        for row in range(1, ws.max_row)
        if ws.cell(row=row + 1, column=2).value
        and ws.cell(row=row + 1, column=9).value
        and ws.cell(row=row + 1, column=2).value != ws.cell(row=row + 1, column=9).value
    ]
    data = copy.deepcopy(data) + [to_exchange(obj) for obj in data]

    # Water unit changes
    data.extend(
        [
            (
                ("Water", ("air",), "kilogram", "biosphere"),
                {"unit": "cubic meter", "multiplier": 0.001},
            ),
            (
                (
                    "Water",
                    ("air", "non-urban air or from high stacks"),
                    "kilogram",
                    "biosphere",
                ),
                {"unit": "cubic meter", "multiplier": 0.001},
            ),
            (
                (
                    "Water",
                    ("air", "lower stratosphere + upper troposphere"),
                    "kilogram",
                    "biosphere",
                ),
                {"unit": "cubic meter", "multiplier": 0.001},
            ),
            (
                (
                    "Water",
                    ("air", "urban air close to ground"),
                    "kilogram",
                    "biosphere",
                ),
                {"unit": "cubic meter", "multiplier": 0.001},
            ),
        ]
    )

    return {"fields": ["name", "categories", "unit", "type"], "data": data}


def get_simapro_water_migration_data():
    return json.load(open(dirpath / "simapro-water.json"))


def get_us_lci_migration_data():
    """Fix US LCI database name inconsistencies"""
    return {
        "fields": ["name"],
        "data": [
            ((k,), {"name": v})
            for k, v in json.load(
                open(dirpath / "us-lci.json", encoding="utf-8")
            ).items()
        ],
    }


def get_exiobase_biosphere_migration_data():
    """Migrate to ecoinvent3 flow names"""
    return json.load(open(dirpath / "exiomigration.json", encoding="utf-8"))


def convert_simapro_ecoinvent_elementary_flows():
    """Write a correspondence list from SimaPro elementary flow names to ecoinvent 3 flow names to a JSON file.

    Uses custom SimaPro specific data. Ecoinvent 2 -> 3 conversion is in a separate JSON file."""
    ws = get_sheet(dirpath / "lci" / "SimaPro - ecoinvent - biosphere.xlsx", "ee")
    data = [
        [ws.cell(row=row + 1, column=col + 1).value for col in range(3)]
        for row in range(1, ws.max_row)
    ]
    data = {(SIMAPRO_BIOSPHERE[obj[0]], obj[1], obj[2]) for obj in data}
    write_json_file(sorted(data), "simapro-biosphere")


def convert_simapro_ecoinvent_3_migration_data():
    VERSIONS = (
        ("Mapping", "3.1"),
        ("Mapping 3.2", "3.2"),
        ("Mapping 3.3", "3.3"),
        ("Mapping 3.4", "3.4"),
        ("Mapping 3.5", "3.5"),
    )

    for ws_name, version in VERSIONS:
        ws = get_sheet(
            dirpath / "lci" / "SimaPro - ecoinvent - technosphere.xlsx", ws_name
        )
        data = [[ws.cell(row=row+1, column=col+1).value for col in range(1, 6)]
                 for row in range(3, ws.max_row)]
        fp = os.path.join(
            dirpath,
            'lci',
            'Simapro - ecoinvent {} mapping.gzip'.format(version)
        )
        with gzip.GzipFile(fp, "w") as fout:
            fout.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))


def get_simapro_ecoinvent_3_migration_data(version):
    """Write a migrations data file from SimaPro activity names to ecoinvent 3 processes.

    Correspondence file is processed from Pré, and has the following fields:

        #. SimaPro name
        #. Ecoinvent flow name
        #. Location
        #. Ecoinvent activity name
        #. System model
        #. SimaPro type

    Note that even the official matching data from Pré is incorrect, but works if we cast all strings to lower case.

    SimaPro type is either ``System terminated`` or ``Unit process``. We always match to unit processes regardless of SimaPro type."""
    fp = dirpath / "lci" / ("Simapro - ecoinvent {} mapping.gzip".format(version))
    with gzip.GzipFile(fp, "r") as fout:
        data = json.loads(fout.read().decode("utf-8"))
    return {
        "fields": ["name"],
        "data": [
            (
                (line[0],),
                {
                    "location": line[2],
                    "name": line[3],
                    "reference product": line[1],
                    "system model": line[4],
                    "simapro name": line[0],
                },
            )
            for line in data
        ],
    }


def convert_ecoinvent_2_301():
    """Write a migrations data file from ecoinvent 2 to 3.1.

    This is not simple, unfortunately. We have to deal with at least the following:
        * Unit changes (e.g. cubic meters to MJ)
        * Some datasets are deleted, and replaced by others

    """
    ws = get_sheet(
        dirpath / "lci" / "ecoinvent 2-3.01.xlsx", "correspondence sheet_corrected"
    )
    data = [
        [ws.cell(row=row + 1, column=col + 1).value for col in range(17)]
        for row in range(1, ws.max_row)
    ]
    data = {
        "fields": ["name", "location"],
        "data": [
            (
                {"name": line[0]},
                {
                    "location": line[2],
                    "name": line[3],
                    "reference product": line[1],
                    "system model": line[4],
                },
            )
            for line in data
        ],
    }
    write_json_file(data, "simapro-ecoinvent31")


def _add_new_ecoinvent_biosphere_flows(version):
    flows = json.load(
        open(dirpath / "lci" / ("ecoinvent {} new biosphere.json".format(version)))
    )

    db = Database(config.biosphere)
    count = 0

    for flow in flows:
        flow["categories"] = tuple(flow["categories"])
        if (config.biosphere, flow["code"]) not in db:
            count += 1
            db.new_activity(**flow).save()

    print("Added {} new biosphere flows".format(count))
    return db


add_ecoinvent_33_biosphere_flows = partial(
    _add_new_ecoinvent_biosphere_flows, version="33"
)
add_ecoinvent_34_biosphere_flows = partial(
    _add_new_ecoinvent_biosphere_flows, version="34"
)
add_ecoinvent_35_biosphere_flows = partial(
    _add_new_ecoinvent_biosphere_flows, version="35"
)
add_ecoinvent_36_biosphere_flows = partial(
    _add_new_ecoinvent_biosphere_flows, version="36"
)
add_ecoinvent_37_biosphere_flows = partial(
    _add_new_ecoinvent_biosphere_flows, version="37"
)
add_ecoinvent_38_biosphere_flows = partial(
    _add_new_ecoinvent_biosphere_flows, version="38"
)


def convert_lcia_methods_data():
    csv_file = csv.reader(
        open(dirpath / "lcia" / "categoryUUIDs.csv", encoding="latin-1"), delimiter=";"
    )
    next(csv_file)  # Skip header row
    csv_data = [
        {
            "name": (line[0], line[2], line[4]),
            # 'unit': line[6],
            "description": line[7],
        }
        for line in csv_file
    ]

    filename = "LCIA_Implementation_3.8.xlsx"
    sheet = get_sheet(dirpath / "lcia" / filename, "CFs")

    def process_row(row):
        data = [cell.value for i, cell in zip(range(8), row)]
        if not isinstance(data[-1], Number):
            return None
        else:
            return {
                "method": tuple(data[:3]),
                "name": data[3],
                "categories": tuple(data[4:6]),
                "amount": data[6],
            }

    cf_data = [process_row(row) for rowidx, row in enumerate(sheet.rows) if rowidx]

    sheet = get_sheet(dirpath / "lcia" / filename, "Indicators")

    def process_unit_row(row):
        data = [cell.value for i, cell in zip(range(4), row)]
        return tuple(data[:3]), data[3]

    units = dict(
        process_unit_row(row) for rowidx, row in enumerate(sheet.rows) if rowidx
    )

    return csv_data, cf_data, units, filename


def get_valid_geonames():
    """Get list of short location names used in ecoinvent 3"""
    return json.load(open(dirpath / "lci" / "geodata.json", encoding="utf-8"))["names"]


def get_ecoinvent_pre35_migration_data():
    return json.load(open(dirpath / "lci" / "ecoinvent_pre35_migration.json"))


def update_db_ecoinvent_locations(database_name):
    """Update ecoinvent location names for an existing database.

    Returns number of modified datasets."""
    from ..strategies.locations import GEO_UPDATE

    db = Database(database_name)
    if not len(db):
        return 0

    count = 0
    for ds in db:
        if ds["location"] in GEO_UPDATE:
            count += 1
            ds["location"] = GEO_UPDATE[ds["location"]]
            ds.save()

    return count


def add_example_database(overwrite=True):
    from ..importers.excel import (
        assign_only_product_as_production,
        convert_activity_parameters_to_list,
        convert_uncertainty_types_to_integers,
        csv_add_missing_exchanges_section,
        csv_drop_unknown,
        csv_numerize,
        csv_restore_booleans,
        csv_restore_tuples,
        drop_falsey_uncertainty_fields_but_keep_zeros,
        ExcelImporter,
        set_code_by_activity_hash,
        strip_biosphere_exc_locations,
    )

    if "Mobility example" in databases:
        if not overwrite:
            print("Example already imported, use `overwrite=True` to delete")
            return
        else:
            del databases["Mobility example"]
            if ("IPCC", "simple") in methods:
                del methods[("IPCC", "simple")]

    importer = ExcelImporter(
        dirpath / "examples" / "sample_parameterized_database.xlsx"
    )
    importer.strategies = [
        csv_restore_tuples,
        csv_restore_booleans,
        csv_numerize,
        csv_drop_unknown,
        csv_add_missing_exchanges_section,
        strip_biosphere_exc_locations,
        set_code_by_activity_hash,
        assign_only_product_as_production,
        drop_falsey_uncertainty_fields_but_keep_zeros,
        convert_uncertainty_types_to_integers,
        convert_activity_parameters_to_list,
    ]
    importer.apply_strategies()
    importer.match_database(fields=["name"])
    importer.write_database(activate_parameters=True)

    group = "Mobility exchanges"
    Group.delete().where(Group.name == group).execute()
    group = Group.create(name=group)

    for ds in Database("Mobility example"):
        parameters.add_exchanges_to_group(group, ds)

    parameters.recalculate()

    ipcc = Method(("IPCC", "simple"))
    ipcc.register()
    ipcc.write([(("Mobility example", "CO2"), 1)])
